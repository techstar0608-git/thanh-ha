#!/usr/bin/env python3
"""
Tạo file Excel template cho website HP và một vài ảnh sản phẩm mẫu.

Sinh ra:
  public/data/san-pham-template.xlsx   -> Template để gửi cho khách điền (luôn tạo mới)
  public/data/san-pham.xlsx            -> File website ĐANG ĐỌC (chỉ tạo nếu chưa có)
  public/sample/*.png                  -> Ảnh sản phẩm mẫu (để demo chạy được ngay)

Chạy:  python3 scripts/make_template.py
"""

import os
import shutil

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(ROOT, "public", "data")
SAMPLE_DIR = os.path.join(ROOT, "public", "sample")
TEMPLATE_PATH = os.path.join(DATA_DIR, "san-pham-template.xlsx")
LIVE_PATH = os.path.join(DATA_DIR, "san-pham.xlsx")

# Màu thương hiệu
GREEN = "4E8A3C"
GREEN_DARK = "356B2A"
GREEN_LIGHT = "EAF3E2"
ORANGE = "E08218"

# ----- Định nghĩa cột (tên cột PHẢI giữ nguyên để web đọc đúng) -----
COLUMNS = [
    ("Mã sản phẩm", 16, "Bắt buộc",
     "Mã định danh, chữ thường không dấu, dùng cho đường dẫn web.", "npk-20-20-15"),
    ("Tên sản phẩm", 28, "Bắt buộc",
     "Tên đầy đủ hiển thị của sản phẩm.", "Phân bón NPK 20-20-15 + TE"),
    ("Loại", 22, "Nên có",
     "Tên nhóm hiển thị trên web.", "Phân bón rễ (NPK)"),
    ("Mã loại", 12, "Nên có",
     "Mã nhóm ngắn để gom sản phẩm cùng danh mục (web KHÔNG hiển thị mã, chỉ dùng "
     "để nhóm/lọc). Cùng mã = cùng nhóm. Xem bảng mã ở sheet Hướng dẫn.", "NPK"),
    ("Mô tả ngắn", 34, "Nên có",
     "Một câu mô tả/slogan ngắn.", "Cân đối đa lượng cho giai đoạn nuôi trái."),
    ("Mô tả chi tiết", 40, "Tùy chọn",
     "Đoạn giới thiệu chi tiết. Xuống dòng bằng Alt+Enter.", "Sản phẩm cung cấp..."),
    ("Thành phần", 34, "Tùy chọn",
     "Mỗi thành phần 1 dòng dạng  Tên = Hàm lượng,  cách nhau bằng ; hoặc xuống dòng.",
     "Đạm (N)=20%; Lân (P2O5)=20%; Kali (K2O)=15%"),
    ("Công dụng", 34, "Nên có",
     "Mỗi công dụng cách nhau bằng ; hoặc xuống dòng.",
     "Thúc đẩy sinh trưởng; Nuôi trái lớn nhanh; Tăng đề kháng"),
    ("Đối tượng cây trồng", 26, "Tùy chọn",
     "Các loại cây dùng được.", "Lúa, cà phê, cây ăn trái, rau màu"),
    ("Cách dùng & liều lượng", 34, "Tùy chọn",
     "Hướng dẫn liều lượng. Xuống dòng bằng Alt+Enter.", "Bón gốc 0.3–0.5 kg/gốc..."),
    ("Quy cách", 18, "Tùy chọn", "Quy cách đóng gói.", "Bao 25kg / 50kg"),
    ("Xuất xứ", 16, "Tùy chọn", "Nơi sản xuất.", "Việt Nam"),
    ("Giá", 16, "Tùy chọn", "Để trống = hiển thị 'Liên hệ'.", "Liên hệ"),
    ("Ảnh chính (link Drive)", 36, "Nên có",
     "Link chia sẻ Google Drive (đặt quyền 'Bất kỳ ai có link').",
     "https://drive.google.com/file/d/FILE_ID/view?usp=sharing"),
    ("Ảnh phụ (link Drive)", 36, "Tùy chọn",
     "Nhiều link Drive, cách nhau bằng ; (thư viện ảnh).",
     "https://drive.google.com/file/d/ID2/view; https://drive.google.com/file/d/ID3/view"),
    ("Nổi bật", 12, "Tùy chọn", "Chọn 'Có' để hiện ở mục nổi bật trang chủ.", "Có"),
]

# ----- Bảng mã loại: (Mã, Tên loại hiển thị) -----
# Mã dùng để GOM NHÓM sản phẩm (web không hiển thị mã). Cùng mã = cùng danh mục.
LOAI_CODES = [
    ("NPK", "Phân bón rễ (NPK)"),
    ("BL", "Phân bón lá"),
    ("HC", "Phân hữu cơ vi sinh"),
    ("TVL", "Phân trung vi lượng"),
    ("DON", "Phân đơn (Đạm/Lân/Kali)"),
    ("SH", "Chế phẩm sinh học"),
]
MA_GOI_Y = [code for code, _ in LOAI_CODES]
LOAI_GOI_Y = [ten for _, ten in LOAI_CODES]

# ----- Dữ liệu mẫu (ảnh dùng file local để demo chạy ngay; khi dùng thật hãy thay bằng link Drive) -----
SAMPLE_ROWS = [
    {
        "Mã sản phẩm": "npk-20-20-15",
        "Tên sản phẩm": "Phân bón NPK 20-20-15 + TE",
        "Loại": "Phân bón rễ (NPK)",
        "Mã loại": "NPK",
        "Mô tả ngắn": "Cân đối đa lượng cho giai đoạn sinh trưởng và nuôi trái.",
        "Mô tả chi tiết": "Công thức NPK cân đối bổ sung vi lượng (TE), giúp cây phát "
                          "triển toàn diện, tăng tỉ lệ đậu trái và chất lượng nông sản.",
        "Thành phần": "Đạm (N)=20%; Lân hữu hiệu (P2O5)=20%; Kali (K2O)=15%; "
                      "Vi lượng (TE)=Bổ sung",
        "Công dụng": "Thúc đẩy sinh trưởng mạnh; Nuôi trái lớn nhanh, chắc hạt; "
                     "Tăng sức đề kháng cho cây",
        "Đối tượng cây trồng": "Lúa, cà phê, cây ăn trái, rau màu",
        "Cách dùng & liều lượng": "Bón gốc 0.3–0.5 kg/gốc tùy tuổi cây.\n"
                                  "Lúa: 8–12 kg/sào/lần, chia 2–3 lần/vụ.",
        "Quy cách": "Bao 25kg / 50kg",
        "Xuất xứ": "Việt Nam",
        "Giá": "Liên hệ",
        "Ảnh chính (link Drive)": "/sample/npk.png",
        "Ảnh phụ (link Drive)": "",
        "Nổi bật": "Có",
    },
    {
        "Mã sản phẩm": "huu-co-vi-sinh",
        "Tên sản phẩm": "Phân hữu cơ vi sinh HP-Organic",
        "Loại": "Phân hữu cơ vi sinh",
        "Mã loại": "HC",
        "Mô tả ngắn": "Cải tạo đất, bổ sung vi sinh vật có lợi cho bộ rễ.",
        "Mô tả chi tiết": "Phân hữu cơ giàu mùn kết hợp vi sinh vật có ích, giúp cải "
                          "tạo đất tơi xốp, phục hồi hệ rễ và tăng hiệu quả hấp thu dinh dưỡng.",
        "Thành phần": "Chất hữu cơ=22%; Acid Humic=2%; Vi sinh vật có ích=1x10^6 CFU/g",
        "Công dụng": "Cải tạo đất tơi xốp; Tăng vi sinh vật có lợi; "
                     "Phục hồi rễ sau ngập úng",
        "Đối tượng cây trồng": "Mọi loại cây trồng",
        "Cách dùng & liều lượng": "Bón lót 1–2 kg/gốc hoặc 200–400 kg/ha.",
        "Quy cách": "Bao 20kg / 40kg",
        "Xuất xứ": "Việt Nam",
        "Giá": "Liên hệ",
        "Ảnh chính (link Drive)": "/sample/huuco.png",
        "Ảnh phụ (link Drive)": "",
        "Nổi bật": "Có",
    },
    {
        "Mã sản phẩm": "bon-la-amino",
        "Tên sản phẩm": "Phân bón lá Amino HP",
        "Loại": "Phân bón lá",
        "Mã loại": "BL",
        "Mô tả ngắn": "Bổ sung amino acid, dưỡng lá xanh dày, mau bén rễ.",
        "Mô tả chi tiết": "Phân bón lá giàu amino acid và vi lượng, hấp thu nhanh "
                          "qua lá, giúp cây phục hồi nhanh, lá xanh dày, tăng quang hợp.",
        "Thành phần": "Amino acid=10%; Đạm (N)=8%; Vi lượng chelate=Bổ sung",
        "Công dụng": "Dưỡng lá xanh dày; Phục hồi cây nhanh; Tăng quang hợp",
        "Đối tượng cây trồng": "Rau màu, cây ăn trái, hoa cây cảnh",
        "Cách dùng & liều lượng": "Pha 20–30 ml/16L nước, phun đều 2 mặt lá, "
                                  "định kỳ 7–10 ngày/lần.",
        "Quy cách": "Chai 500ml / Can 5L",
        "Xuất xứ": "Việt Nam",
        "Giá": "Liên hệ",
        "Ảnh chính (link Drive)": "/sample/bonla.png",
        "Ảnh phụ (link Drive)": "",
        "Nổi bật": "",
    },
    {
        "Mã sản phẩm": "canxi-bo",
        "Tên sản phẩm": "Trung vi lượng Canxi - Bo",
        "Loại": "Phân trung vi lượng",
        "Mã loại": "TVL",
        "Mô tả ngắn": "Chống nứt trái, rụng hoa, tăng đậu quả.",
        "Mô tả chi tiết": "Bổ sung Canxi và Bo dễ hấp thu, giúp cứng cây, hạn chế "
                          "nứt trái – rụng hoa, tăng tỉ lệ đậu quả và độ bền trái.",
        "Thành phần": "Canxi (CaO)=10%; Bo (B)=0.5%",
        "Công dụng": "Chống nứt trái; Giảm rụng hoa, rụng trái non; Tăng đậu quả",
        "Đối tượng cây trồng": "Cây ăn trái, cà chua, ớt, dưa",
        "Cách dùng & liều lượng": "Pha 25–30 ml/16L nước, phun giai đoạn ra hoa – nuôi trái.",
        "Quy cách": "Chai 500ml",
        "Xuất xứ": "Việt Nam",
        "Giá": "Liên hệ",
        "Ảnh chính (link Drive)": "/sample/canxibo.png",
        "Ảnh phụ (link Drive)": "",
        "Nổi bật": "",
    },
]


def make_sample_images():
    """Tạo ảnh sản phẩm mẫu on-brand để demo hiển thị được ngay."""
    from PIL import Image, ImageDraw, ImageFont

    os.makedirs(SAMPLE_DIR, exist_ok=True)
    font_path = None
    for cand in [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/Library/Fonts/Arial.ttf",
        "/System/Library/Fonts/Helvetica.ttc",
    ]:
        if os.path.exists(cand):
            font_path = cand
            break

    def font(size):
        return ImageFont.truetype(font_path, size) if font_path else ImageFont.load_default()

    items = [
        ("npk.png", "NPK\n20-20-15", (78, 138, 60)),
        ("huuco.png", "HỮU CƠ\nVI SINH", (53, 107, 42)),
        ("bonla.png", "BÓN LÁ\nAMINO", (107, 163, 74)),
        ("canxibo.png", "CANXI\nBO", (224, 130, 24)),
    ]
    W = H = 800
    for name, label, (r, g, b) in items:
        img = Image.new("RGB", (W, H), (245, 250, 240))
        d = ImageDraw.Draw(img)
        # nền gradient nhẹ
        for y in range(H):
            t = y / H
            d.line(
                [(0, y), (W, y)],
                fill=(int(245 - t * 30), int(250 - t * 20), int(240 - t * 30)),
            )
        # "chai/bao" tròn ở giữa
        d.ellipse([180, 180, 620, 620], fill=(r, g, b))
        d.ellipse([180, 180, 620, 620], outline=(255, 255, 255), width=10)
        f = font(86)
        lines = label.split("\n")
        total_h = sum(
            d.textbbox((0, 0), ln, font=f)[3] - d.textbbox((0, 0), ln, font=f)[1] + 14
            for ln in lines
        )
        y = (H - total_h) / 2
        for ln in lines:
            bb = d.textbbox((0, 0), ln, font=f)
            w = bb[2] - bb[0]
            d.text(((W - w) / 2, y), ln, font=f, fill=(255, 255, 255))
            y += (bb[3] - bb[1]) + 14
        d.text((W / 2 - 28, 700), "HP", font=font(48), fill=(r, g, b))
        img.save(os.path.join(SAMPLE_DIR, name))
    print(f"✓ Đã tạo {len(items)} ảnh mẫu trong public/sample/")


def style_header(ws, row, cols, fill=GREEN, color="FFFFFF"):
    thin = Side(style="thin", color="D6E2D0")
    for i, _ in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i)
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(bold=True, color=color, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_huongdan(wb):
    ws = wb.create_sheet("Hướng dẫn")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 46

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "HƯỚNG DẪN NHẬP SẢN PHẨM — Website HP"
    t.font = Font(bold=True, size=16, color=GREEN_DARK)
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    intro = [
        "• Nhập sản phẩm tại sheet 'SanPham' — MỖI DÒNG LÀ MỘT SẢN PHẨM.",
        "• KHÔNG đổi tên các cột ở dòng tiêu đề (web đọc theo tên cột).",
        "• Có thể xoá các dòng mẫu và nhập dữ liệu của bạn.",
        "• Lưu file rồi gửi lại — đổi tên thành 'san-pham.xlsx' đặt vào thư mục public/data.",
    ]
    r = 3
    for line in intro:
        ws.cell(row=r, column=1, value=line).font = Font(size=11)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Ý NGHĨA CÁC CỘT").font = Font(bold=True, size=12, color=GREEN_DARK)
    r += 1
    head = ["Tên cột", "Bắt buộc", "Mô tả", "Ví dụ"]
    for i, h in enumerate(head, start=1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, head)
    r += 1
    thin = Side(style="thin", color="E3EBE5")
    for name, _w, req, desc, ex in COLUMNS:
        ws.cell(row=r, column=1, value=name).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=req)
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=4, value=ex)
        for i in range(1, 5):
            cell = ws.cell(row=r, column=i)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if req == "Bắt buộc":
                ws.cell(row=r, column=2).font = Font(bold=True, color="C0392B")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="BẢNG MÃ LOẠI (để gom nhóm sản phẩm)").font = Font(
        bold=True, size=12, color=GREEN_DARK
    )
    r += 1
    ws.cell(
        row=r, column=1,
        value="Điền 'Mã loại' theo bảng dưới. Web KHÔNG hiển thị mã — chỉ dùng để gom "
        "các sản phẩm cùng mã vào một danh mục. Có thể thêm mã mới của bạn.",
    ).font = Font(size=11)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    code_head = ["Mã loại", "Tên loại hiển thị"]
    for i, h in enumerate(code_head, start=1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, code_head)
    r += 1
    for code, ten in LOAI_CODES:
        ws.cell(row=r, column=1, value=code).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=ten)
        for i in range(1, 3):
            cell = ws.cell(row=r, column=i)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CÁCH LẤY LINK ẢNH TỪ GOOGLE DRIVE").font = Font(
        bold=True, size=12, color=ORANGE
    )
    r += 1
    drive_steps = [
        "1. Tải ảnh sản phẩm lên Google Drive.",
        "2. Chuột phải vào ảnh → Chia sẻ → 'Bất kỳ ai có đường liên kết' (Anyone with the link).",
        "3. Bấm 'Sao chép liên kết' và dán vào cột 'Ảnh chính (link Drive)'.",
        "   Ví dụ: https://drive.google.com/file/d/1AbCdEf.../view?usp=sharing",
        "4. Nhiều ảnh phụ: dán nhiều link vào cột 'Ảnh phụ', cách nhau bằng dấu ;",
        "Lưu ý: ảnh PHẢI ở chế độ công khai (public) thì web mới hiển thị được.",
    ]
    for line in drive_steps:
        ws.cell(row=r, column=1, value=line).font = Font(size=11)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
    return ws


def build_sanpham(wb):
    ws = wb.active
    ws.title = "SanPham"
    ws.sheet_view.showGridLines = False
    names = [c[0] for c in COLUMNS]

    # tiêu đề
    for i, name in enumerate(names, start=1):
        ws.cell(row=1, column=i, value=name)
        ws.column_dimensions[get_column_letter(i)].width = COLUMNS[i - 1][1]
    style_header(ws, 1, names, fill=GREEN)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"

    # dữ liệu mẫu
    thin = Side(style="thin", color="EDF2EC")
    for ridx, row in enumerate(SAMPLE_ROWS, start=2):
        for cidx, name in enumerate(names, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=row.get(name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if ridx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="FAFCF8")
        ws.row_dimensions[ridx].height = 64

    last_row = 400  # áp validation cho nhiều dòng trống phía dưới

    # Dropdown gợi ý cho cột "Loại" (không bắt buộc đúng -> cho phép nhập tự do)
    loai_col = get_column_letter(names.index("Loại") + 1)
    dv_loai = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(LOAI_GOI_Y),
        allow_blank=True,
        showErrorMessage=False,
    )
    dv_loai.prompt = "Chọn hoặc tự nhập nhóm sản phẩm"
    dv_loai.promptTitle = "Loại sản phẩm"
    ws.add_data_validation(dv_loai)
    dv_loai.add(f"{loai_col}2:{loai_col}{last_row}")

    # Dropdown mã loại — chọn mã ngắn để gom nhóm (cho phép tự nhập mã mới)
    maloai_col = get_column_letter(names.index("Mã loại") + 1)
    dv_maloai = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(MA_GOI_Y),
        allow_blank=True,
        showErrorMessage=False,
    )
    dv_maloai.prompt = "Chọn/nhập mã nhóm — cùng mã sẽ được gom chung danh mục"
    dv_maloai.promptTitle = "Mã loại"
    ws.add_data_validation(dv_maloai)
    dv_maloai.add(f"{maloai_col}2:{maloai_col}{last_row}")

    # Dropdown Có/Không cho cột "Nổi bật"
    nb_col = get_column_letter(names.index("Nổi bật") + 1)
    dv_nb = DataValidation(
        type="list", formula1='"Có,Không"', allow_blank=True, showErrorMessage=False
    )
    ws.add_data_validation(dv_nb)
    dv_nb.add(f"{nb_col}2:{nb_col}{last_row}")
    return ws


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    make_sample_images()

    wb = Workbook()
    build_sanpham(wb)
    build_huongdan(wb)
    # đưa sheet Hướng dẫn lên đầu
    wb.move_sheet("Hướng dẫn", -1)
    wb.active = wb["SanPham"]

    wb.save(TEMPLATE_PATH)
    print(f"✓ Đã tạo template: {os.path.relpath(TEMPLATE_PATH, ROOT)}")

    if not os.path.exists(LIVE_PATH):
        shutil.copyfile(TEMPLATE_PATH, LIVE_PATH)
        print(f"✓ Đã tạo file dữ liệu web: {os.path.relpath(LIVE_PATH, ROOT)}")
    else:
        print(f"• Giữ nguyên file dữ liệu web hiện có: {os.path.relpath(LIVE_PATH, ROOT)}")
        print("  (Xoá file này rồi chạy lại nếu muốn ghi đè bằng dữ liệu mẫu.)")


if __name__ == "__main__":
    main()
